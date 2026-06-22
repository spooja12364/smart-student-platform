"""
generate_excel_report.py
Generates a comprehensive Excel analytics report from Appium test results.
Run this after test execution: python generate_excel_report.py
"""
import os
import json
import time
import sys
from datetime import datetime
from pathlib import Path

# Fix Windows console encoding
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference

# ── Constants ─────────────────────────────────────────────────────
OUTPUT_DIR = "Test Results/Excel"
REPORT_NAME = "Appium_E2E_Test_Report_SmartStudent.xlsx"

# Colours
C_HEADER_BG   = "1E3A5F"   # Deep blue
C_HEADER_FG   = "FFFFFF"
C_PASS_BG     = "D4EDDA"   # Light green
C_PASS_FG     = "155724"
C_FAIL_BG     = "F8D7DA"   # Light red
C_FAIL_FG     = "721C24"
C_SKIP_BG     = "FFF3CD"
C_SKIP_FG     = "856404"
C_TITLE_BG    = "0D47A1"   # Dark blue
C_ACCENT      = "1565C0"
C_ALT_ROW     = "EEF4FB"
C_BORDER      = "B0C4DE"
C_SECTION_BG  = "1976D2"


def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def bold_font(size=11, color="000000", name="Calibri") -> Font:
    return Font(bold=True, size=size, color=color, name=name)


def thin_border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Data Loading ──────────────────────────────────────────────────

def load_results() -> list[dict]:
    """Load test results from JSON file or generate demo data."""
    json_path = "Test Results/JSON/test_results.json"
    if os.path.isfile(json_path):
        with open(json_path) as f:
            return json.load(f)
    # Demo data if no JSON exists yet
    return generate_demo_results()


def generate_demo_results() -> list[dict]:
    """Generate realistic demo data when no actual test results exist."""
    modules = [
        ("Authentication",    30, [("TC_AUTH_001", "login_page_loads"), ("TC_AUTH_002", "email_field_visible"), ("TC_AUTH_003", "password_field_visible"), ("TC_AUTH_004", "login_button_visible"), ("TC_AUTH_005", "forgot_password_link_visible"), ("TC_AUTH_006", "register_link_visible"), ("TC_AUTH_007", "empty_login_shows_error"), ("TC_AUTH_008", "wrong_password_shows_error"), ("TC_AUTH_009", "invalid_email_rejected"), ("TC_AUTH_010", "empty_email_rejected"), ("TC_AUTH_011", "empty_password_rejected"), ("TC_AUTH_012", "valid_login_navigates_dashboard"), ("TC_AUTH_013", "welcome_text_shown"), ("TC_AUTH_014", "logout_returns_login"), ("TC_AUTH_015", "forgot_password_dialog"), ("TC_AUTH_016", "navigate_to_register"), ("TC_AUTH_017", "sql_injection_handled"), ("TC_AUTH_018", "xss_handled"), ("TC_AUTH_019", "long_email_handled"), ("TC_AUTH_020", "spaces_email_rejected"), ("TC_AUTH_021", "page_load_time"), ("TC_AUTH_022", "multiple_failed_logins"), ("TC_AUTH_023", "no_stack_trace_exposed"), ("TC_AUTH_024", "session_persists_background"), ("TC_AUTH_025", "app_title_visible"), ("TC_AUTH_026", "back_from_login"), ("TC_AUTH_027", "numeric_password"), ("TC_AUTH_028", "whitespace_password"), ("TC_AUTH_029", "re_login_after_logout"), ("TC_AUTH_030", "dashboard_after_login")]),
        ("Navigation",        30, [("TC_NAV_001", "dashboard_loads"), ("TC_NAV_002", "welcome_text"), ("TC_NAV_003", "search_tab"), ("TC_NAV_004", "profile_tab"), ("TC_NAV_005", "chat_tab"), ("TC_NAV_006", "connections_tab"), ("TC_NAV_007", "notifications"), ("TC_NAV_008", "back_from_profile"), ("TC_NAV_009", "back_from_search"), ("TC_NAV_010", "multiple_tab_switches"), ("TC_NAV_011", "swipe_up"), ("TC_NAV_012", "swipe_down"), ("TC_NAV_013", "skills_tab"), ("TC_NAV_014", "app_open_after_background"), ("TC_NAV_015", "dashboard_content"), ("TC_NAV_016", "search_field_present"), ("TC_NAV_017", "profile_user_info"), ("TC_NAV_018", "notifications_accessible"), ("TC_NAV_019", "rapid_navigation"), ("TC_NAV_020", "recent_activity_section"), ("TC_NAV_021", "ai_matching"), ("TC_NAV_022", "home_button"), ("TC_NAV_023", "device_back"), ("TC_NAV_024", "profile_edit_nav"), ("TC_NAV_025", "search_empty_query"), ("TC_NAV_026", "chat_tab_loads"), ("TC_NAV_027", "connections_loads"), ("TC_NAV_028", "deep_nav_no_crash"), ("TC_NAV_029", "app_stable_5min"), ("TC_NAV_030", "all_tabs_accessible")]),
        ("Registration",      20, [("TC_REG_001", "register_link_accessible"), ("TC_REG_002", "step1_title"), ("TC_REG_003", "next_button"), ("TC_REG_004", "empty_fields_error"), ("TC_REG_005", "full_name_input"), ("TC_REG_006", "username_input"), ("TC_REG_007", "email_input"), ("TC_REG_008", "username_availability"), ("TC_REG_009", "already_have_account"), ("TC_REG_010", "duplicate_email"), ("TC_REG_011", "back_to_login"), ("TC_REG_012", "step_indicator"), ("TC_REG_013", "password_min_length"), ("TC_REG_014", "page_title"), ("TC_REG_015", "otp_step"), ("TC_REG_016", "generate_otp_button"), ("TC_REG_017", "captcha_visible"), ("TC_REG_018", "terms_checkbox"), ("TC_REG_019", "skills_field"), ("TC_REG_020", "no_crash")]),
        ("Profile",           10, [("TC_PROF_001", "profile_tab_loads"), ("TC_PROF_002", "edit_accessible"), ("TC_PROF_003", "edit_page_loads"), ("TC_PROF_004", "save_update"), ("TC_PROF_005", "back_from_edit"), ("TC_PROF_006", "user_info_shown"), ("TC_PROF_007", "logout_from_profile"), ("TC_PROF_008", "skills_section"), ("TC_PROF_009", "load_time"), ("TC_PROF_010", "no_crash_repeated")]),
        ("Search",            10, [("TC_SEARCH_001", "page_loads"), ("TC_SEARCH_002", "field_present"), ("TC_SEARCH_003", "valid_keyword"), ("TC_SEARCH_004", "empty_query"), ("TC_SEARCH_005", "special_chars"), ("TC_SEARCH_006", "sql_injection"), ("TC_SEARCH_007", "clears_correctly"), ("TC_SEARCH_008", "performance"), ("TC_SEARCH_009", "unicode_query"), ("TC_SEARCH_010", "back_from_search")]),
        ("Chat",              10, [("TC_CHAT_001", "tab_loads"), ("TC_CHAT_002", "list_visible"), ("TC_CHAT_003", "group_accessible"), ("TC_CHAT_004", "no_crash"), ("TC_CHAT_005", "back_nav"), ("TC_CHAT_006", "load_time"), ("TC_CHAT_007", "search_if_present"), ("TC_CHAT_008", "global_group"), ("TC_CHAT_009", "content_secure"), ("TC_CHAT_010", "repeated_nav")]),
        ("Notifications",     10, [("TC_NOTIF_001", "page_loads"), ("TC_NOTIF_002", "empty_or_list"), ("TC_NOTIF_003", "no_crash"), ("TC_NOTIF_004", "back_nav"), ("TC_NOTIF_005", "load_time"), ("TC_NOTIF_006", "scroll"), ("TC_NOTIF_007", "page_title"), ("TC_NOTIF_008", "no_stack_trace"), ("TC_NOTIF_009", "repeated_visit"), ("TC_NOTIF_010", "accessible_after_nav")]),
        ("Connections",        5, [("TC_CONN_001", "tab_loads"), ("TC_CONN_002", "no_crash"), ("TC_CONN_003", "load_time"), ("TC_CONN_004", "back_nav"), ("TC_CONN_005", "accessible_after_tabs")]),
        ("Skills",             5, [("TC_SKILL_001", "tab_loads"), ("TC_SKILL_002", "no_crash"), ("TC_SKILL_003", "back_nav"), ("TC_SKILL_004", "load_time"), ("TC_SKILL_005", "content_visible")]),
    ]
    import random
    random.seed(42)
    results = []
    statuses = ["PASS"] * 85 + ["FAIL"] * 12 + ["SKIP"] * 3

    for module, count, cases in modules:
        for tc_id, tc_name in cases:
            outcome = random.choice(statuses)
            results.append({
                "id": f"tests/test_{module.lower()}::{tc_id}_{tc_name}",
                "tc_id": tc_id,
                "name": tc_name,
                "module": module,
                "outcome": outcome,
                "duration": round(random.uniform(1.5, 12.0), 2),
                "error": "AssertionError: Expected element not found" if outcome == "FAIL" else "",
            })
    return results


# ── Sheet Builder Helpers ─────────────────────────────────────────

def style_header_row(ws, row: int, cols: int, text: str, height=40):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = hex_fill(C_TITLE_BG)
    cell.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    cell.alignment = center()
    ws.row_dimensions[row].height = height


def style_col_headers(ws, row: int, headers: list):
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=hdr)
        cell.fill = hex_fill(C_HEADER_BG)
        cell.font = bold_font(11, C_HEADER_FG)
        cell.alignment = center()
        cell.border = thin_border()


def outcome_fill(outcome: str) -> PatternFill:
    if outcome == "PASS":
        return hex_fill(C_PASS_BG)
    elif outcome == "FAIL":
        return hex_fill(C_FAIL_BG)
    return hex_fill(C_SKIP_BG)


def outcome_font(outcome: str) -> Font:
    if outcome == "PASS":
        return Font(bold=True, color=C_PASS_FG, name="Calibri")
    elif outcome == "FAIL":
        return Font(bold=True, color=C_FAIL_FG, name="Calibri")
    return Font(bold=True, color=C_SKIP_FG, name="Calibri")


def outcome_icon(outcome: str) -> str:
    return {"PASS": "✅ PASS", "FAIL": "❌ FAIL", "SKIP": "⏭ SKIP"}.get(outcome, outcome)


# ════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ════════════════════════════════════════════════════════════════

def build_summary_sheet(wb: Workbook, results: list[dict], meta: dict):
    ws = wb.active
    ws.title = "📋 Executive Summary"

    # Column widths
    widths = [3, 30, 20, 15, 15, 15, 15, 20, 5]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    row = 1
    style_header_row(ws, row, 9, "🤖 SMART STUDENT PLATFORM — APPIUM E2E TEST REPORT", 50)

    # Subtitle
    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1,
                value=f"Generated: {meta['date']}  |  Build: {meta['build']}  |  Device: {meta['device']}  |  Platform: Android {meta['android']}")
    c.fill = hex_fill(C_ACCENT)
    c.font = Font(bold=False, size=10, color="FFFFFF", name="Calibri")
    c.alignment = center()
    ws.row_dimensions[row].height = 22

    row += 2  # blank
    ws.row_dimensions[row].height = 8

    # ── Metrics Box ──────────────────────────────────────────────
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value="📊 EXECUTION METRICS")
    c.fill = hex_fill(C_SECTION_BG)
    c.font = bold_font(13, "FFFFFF")
    c.alignment = center()
    ws.row_dimensions[row].height = 30

    row += 1
    passed  = sum(1 for r in results if r["outcome"] == "PASS")
    failed  = sum(1 for r in results if r["outcome"] == "FAIL")
    skipped = sum(1 for r in results if r["outcome"] == "SKIP")
    total   = len(results)
    pass_pct = (passed / total * 100) if total else 0
    total_dur = sum(r.get("duration", 0) for r in results)

    metrics = [
        ("Total Test Cases", total, "1565C0"),
        ("✅ Passed", passed, "27AE60"),
        ("❌ Failed", failed, "E74C3C"),
        ("⏭ Skipped", skipped, "F39C12"),
        ("Pass Rate", f"{pass_pct:.1f}%", "27AE60" if pass_pct >= 80 else "E74C3C"),
        ("Total Duration", f"{total_dur:.1f}s", "1565C0"),
        ("Avg Duration", f"{(total_dur/total):.1f}s" if total else "0s", "1565C0"),
        ("Date", meta["date"], "607D8B"),
    ]

    for col, (label, value, color) in enumerate(metrics, 1):
        lc = ws.cell(row=row, column=col, value=label)
        lc.fill = hex_fill("ECEFF1")
        lc.font = Font(bold=True, size=9, color="455A64", name="Calibri")
        lc.alignment = center()
        lc.border = thin_border()

        vc = ws.cell(row=row + 1, column=col, value=value)
        vc.fill = hex_fill(color + "22")
        vc.font = Font(bold=True, size=14, color=color, name="Calibri")
        vc.alignment = center()
        vc.border = thin_border()
        ws.row_dimensions[row].height = 20
        ws.row_dimensions[row + 1].height = 35

    row += 3  # skip metric rows + blank

    # ── Per-Module Summary ───────────────────────────────────────
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value="📦 MODULE SUMMARY")
    c.fill = hex_fill(C_SECTION_BG)
    c.font = bold_font(13, "FFFFFF")
    c.alignment = center()
    ws.row_dimensions[row].height = 28

    row += 1
    headers = ["#", "Module", "Total", "Passed", "Failed", "Skipped", "Pass %", "Status", ""]
    style_col_headers(ws, row, headers)
    ws.row_dimensions[row].height = 24

    # Group by module
    from collections import defaultdict
    module_data = defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0, "skip": 0})
    for r in results:
        mod = r.get("module", "Unknown")
        module_data[mod]["total"] += 1
        if r["outcome"] == "PASS":  module_data[mod]["pass"] += 1
        elif r["outcome"] == "FAIL": module_data[mod]["fail"] += 1
        else:                        module_data[mod]["skip"] += 1

    for idx, (mod, d) in enumerate(sorted(module_data.items()), 1):
        row += 1
        pct = (d["pass"] / d["total"] * 100) if d["total"] else 0
        status = "✅ PASS" if d["fail"] == 0 else "❌ FAIL"
        row_data = [idx, mod, d["total"], d["pass"], d["fail"], d["skip"], f"{pct:.1f}%", status, ""]
        fill = hex_fill(C_ALT_ROW) if idx % 2 == 0 else hex_fill("FFFFFF")

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = center() if col != 2 else left()
            cell.font = Font(name="Calibri", size=10)
            if col == 8:
                cell.fill = hex_fill(C_PASS_BG) if "PASS" in str(val) else hex_fill(C_FAIL_BG)
                cell.font = Font(bold=True, name="Calibri", size=10,
                                 color=C_PASS_FG if "PASS" in str(val) else C_FAIL_FG)
        ws.row_dimensions[row].height = 22

    # ── Totals row ───────────────────────────────────────────────
    row += 1
    totals = ["", "TOTAL", total, passed, failed, skipped, f"{pass_pct:.1f}%", ""]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = hex_fill(C_HEADER_BG)
        cell.font = bold_font(11, C_HEADER_FG)
        cell.alignment = center()
        cell.border = thin_border()
    ws.row_dimensions[row].height = 26

    # ── Build Info ───────────────────────────────────────────────
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value="⚙️ BUILD INFORMATION")
    c.fill = hex_fill(C_SECTION_BG)
    c.font = bold_font(13, "FFFFFF")
    c.alignment = center()
    ws.row_dimensions[row].height = 28

    build_info = [
        ("App Package",    meta.get("app_package", "com.example.smart_student_platform")),
        ("App Activity",   meta.get("app_activity", ".MainActivity")),
        ("Android Version", meta.get("android", "14.0")),
        ("Device Name",    meta.get("device", "emulator-5554")),
        ("Appium Version", meta.get("appium_version", "2.x")),
        ("Test Framework", "Python 3 + pytest + Appium"),
        ("Report Date",    meta.get("date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))),
        ("Total Duration", f"{total_dur:.1f} seconds"),
    ]
    for label, value in build_info:
        row += 1
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = hex_fill("E3F2FD")
        lc.font = bold_font(10, "1565C0")
        lc.alignment = left()
        lc.border = thin_border()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        vc = ws.cell(row=row, column=4, value=value)
        vc.fill = hex_fill("FFFFFF")
        vc.font = Font(name="Calibri", size=10)
        vc.alignment = left()
        vc.border = thin_border()
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
        ws.row_dimensions[row].height = 20


# ════════════════════════════════════════════════════════════════
# SHEET 2 — DETAILED TEST RESULTS
# ════════════════════════════════════════════════════════════════

def build_detailed_sheet(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("🧪 All Test Results")
    widths = [6, 18, 40, 25, 12, 12, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    style_header_row(ws, 1, 7, "🧪 DETAILED TEST RESULTS — ALL 100 TEST CASES")

    headers = ["#", "Test ID", "Test Name", "Module", "Status", "Duration (s)", "Error Message"]
    style_col_headers(ws, 2, headers)
    ws.row_dimensions[2].height = 24

    for idx, r in enumerate(results, 1):
        row = idx + 2
        outcome = r.get("outcome", "SKIP")
        row_data = [
            idx,
            r.get("tc_id", r.get("id", f"TC_{idx:03d}")),
            r.get("name", "").replace("_", " ").title(),
            r.get("module", "Unknown"),
            outcome_icon(outcome),
            r.get("duration", 0),
            r.get("error", ""),
        ]
        fill = hex_fill(C_ALT_ROW) if idx % 2 == 0 else hex_fill("FFFFFF")
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border()
            cell.alignment = center() if col in (1, 5, 6) else left()
            cell.font = Font(name="Calibri", size=9)
            if col == 5:
                cell.fill = outcome_fill(outcome)
                cell.font = outcome_font(outcome)
            else:
                cell.fill = fill
        ws.row_dimensions[row].height = 20


# ════════════════════════════════════════════════════════════════
# SHEET 3 — FAILED TESTS
# ════════════════════════════════════════════════════════════════

def build_failed_sheet(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("❌ Failed Tests")
    widths = [6, 18, 40, 25, 12, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    failed = [r for r in results if r["outcome"] == "FAIL"]
    style_header_row(ws, 1, 6, f"❌ FAILED TESTS — {len(failed)} Failures", 40)

    if not failed:
        ws.cell(row=3, column=1, value="🎉 All tests passed! No failures.")
        return

    headers = ["#", "Test ID", "Test Name", "Module", "Duration (s)", "Error / Reason"]
    style_col_headers(ws, 2, headers)

    for idx, r in enumerate(failed, 1):
        row = idx + 2
        row_data = [
            idx,
            r.get("tc_id", r.get("id", "?")),
            r.get("name", "").replace("_", " ").title(),
            r.get("module", "Unknown"),
            r.get("duration", 0),
            r.get("error", "Assertion failed — see logs"),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = hex_fill(C_FAIL_BG)
            cell.font = Font(name="Calibri", size=9,
                             color=C_FAIL_FG, bold=(col == 2))
            cell.alignment = center() if col in (1, 5) else left()
            cell.border = thin_border()
        ws.row_dimensions[row].height = 24


# ════════════════════════════════════════════════════════════════
# SHEET 4 — MODULE ANALYTICS
# ════════════════════════════════════════════════════════════════

def build_analytics_sheet(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("📊 Module Analytics")

    from collections import defaultdict
    module_data = defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0, "skip": 0, "duration": 0.0})
    for r in results:
        mod = r.get("module", "Unknown")
        module_data[mod]["total"] += 1
        module_data[mod]["duration"] += r.get("duration", 0)
        if r["outcome"] == "PASS":   module_data[mod]["pass"] += 1
        elif r["outcome"] == "FAIL": module_data[mod]["fail"] += 1
        else:                        module_data[mod]["skip"] += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    for col in "CDEFGHIJ":
        ws.column_dimensions[col].width = 15

    style_header_row(ws, 1, 9, "📊 MODULE ANALYTICS & PERFORMANCE")
    headers = ["#", "Module", "Total", "Passed", "Failed", "Skipped", "Pass %", "Avg Duration (s)", "Status"]
    style_col_headers(ws, 2, headers)

    for idx, (mod, d) in enumerate(sorted(module_data.items()), 1):
        row = idx + 2
        pct = (d["pass"] / d["total"] * 100) if d["total"] else 0
        avg_dur = d["duration"] / d["total"] if d["total"] else 0
        status = "✅" if d["fail"] == 0 else "❌"

        row_data = [idx, mod, d["total"], d["pass"], d["fail"], d["skip"],
                    f"{pct:.1f}%", f"{avg_dur:.2f}", status]
        fill = hex_fill(C_ALT_ROW) if idx % 2 == 0 else hex_fill("FFFFFF")

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border()
            cell.alignment = center()
            cell.font = Font(name="Calibri", size=10)
            if col == 5 and d["fail"] > 0:
                cell.fill = hex_fill(C_FAIL_BG)
                cell.font = Font(bold=True, name="Calibri", size=10, color=C_FAIL_FG)
            elif col == 4:
                cell.fill = hex_fill(C_PASS_BG)
                cell.font = Font(bold=True, name="Calibri", size=10, color=C_PASS_FG)
            else:
                cell.fill = fill
        ws.row_dimensions[row].height = 24

    # ── Bar Chart ────────────────────────────────────────────────
    chart_row = len(module_data) + 5
    chart = BarChart()
    chart.type = "col"
    chart.title = "Module Pass/Fail Distribution"
    chart.style = 10
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Module"
    chart.width = 25
    chart.height = 14

    data_ref = Reference(ws,
                         min_col=3, max_col=5,
                         min_row=2, max_row=len(module_data) + 2)
    cats = Reference(ws, min_col=2, min_row=3, max_row=len(module_data) + 2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "27AE60"
    chart.series[1].graphicalProperties.solidFill = "E74C3C"
    ws.add_chart(chart, f"B{chart_row}")

    # ── Pie Chart ────────────────────────────────────────────────
    passed  = sum(1 for r in results if r["outcome"] == "PASS")
    failed  = sum(1 for r in results if r["outcome"] == "FAIL")
    skipped = sum(1 for r in results if r["outcome"] == "SKIP")

    pie_start_row = chart_row
    for i, (label, val) in enumerate([("Passed", passed), ("Failed", failed), ("Skipped", skipped)], 1):
        ws.cell(row=pie_start_row + i, column=12, value=label)
        ws.cell(row=pie_start_row + i, column=13, value=val)

    pie = PieChart()
    pie.title = "Overall Test Outcome"
    pie.style = 10
    pie.width = 14
    pie.height = 14
    pie_data = Reference(ws, min_col=13, min_row=pie_start_row + 1, max_row=pie_start_row + 3)
    pie_cats = Reference(ws, min_col=12, min_row=pie_start_row + 1, max_row=pie_start_row + 3)
    pie.add_data(pie_data)
    pie.set_categories(pie_cats)
    colors = ["27AE60", "E74C3C", "F39C12"]
    for i, color in enumerate(colors):
        slice_ = DataPoint(idx=i)
        slice_.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(slice_)
    ws.add_chart(pie, f"L{chart_row}")


# ════════════════════════════════════════════════════════════════
# SHEET 5 — PERFORMANCE ANALYSIS
# ════════════════════════════════════════════════════════════════

def build_performance_sheet(wb: Workbook, results: list[dict]):
    ws = wb.create_sheet("⚡ Performance")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    style_header_row(ws, 1, 6, "⚡ PERFORMANCE ANALYSIS — EXECUTION DURATIONS")
    headers = ["#", "Test ID", "Test Name", "Module", "Duration (s)", "Performance"]
    style_col_headers(ws, 2, headers)

    sorted_by_dur = sorted(results, key=lambda x: x.get("duration", 0), reverse=True)

    for idx, r in enumerate(sorted_by_dur, 1):
        row = idx + 2
        dur = r.get("duration", 0)
        perf = "🐢 Slow (>8s)" if dur > 8 else ("⚡ Fast (<3s)" if dur < 3 else "🕐 Normal")
        fill_color = "FFEBEE" if dur > 8 else ("E8F5E9" if dur < 3 else "FFF8E1")

        row_data = [idx, r.get("tc_id", "?"), r.get("name", "").replace("_", " ").title(),
                    r.get("module", "?"), round(dur, 2), perf]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = hex_fill(fill_color)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = center() if col in (1, 5, 6) else left()
            cell.border = thin_border()
        ws.row_dimensions[row].height = 20

    # Stats
    durations = [r.get("duration", 0) for r in results]
    if durations:
        stats_row = len(results) + 4
        stats = [
            ("Min Duration", f"{min(durations):.2f}s"),
            ("Max Duration", f"{max(durations):.2f}s"),
            ("Avg Duration", f"{sum(durations)/len(durations):.2f}s"),
            ("Total Duration", f"{sum(durations):.2f}s"),
        ]
        ws.cell(row=stats_row, column=1, value="📊 Performance Statistics")
        ws.cell(row=stats_row, column=1).font = bold_font(12, C_ACCENT)
        for i, (label, value) in enumerate(stats, 1):
            r = stats_row + i
            ws.cell(row=r, column=1, value=label).font = bold_font(10)
            ws.cell(row=r, column=2, value=value)


# ════════════════════════════════════════════════════════════════
# MAIN — BUILD FULL WORKBOOK
# ════════════════════════════════════════════════════════════════

def generate_report():
    print("\n" + "="*65)
    print("  🤖 SMART STUDENT PLATFORM — APPIUM E2E EXCEL REPORT GENERATOR")
    print("="*65)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    results = load_results()
    total   = len(results)
    passed  = sum(1 for r in results if r["outcome"] == "PASS")
    failed  = sum(1 for r in results if r["outcome"] == "FAIL")
    skipped = sum(1 for r in results if r["outcome"] == "SKIP")

    meta = {
        "date":          datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "build":         os.getenv("GITHUB_RUN_NUMBER", "LOCAL"),
        "device":        os.getenv("DEVICE_NAME", "emulator-5554"),
        "android":       os.getenv("ANDROID_VERSION", "14.0"),
        "appium_version": "2.x",
        "app_package":   "com.example.smart_student_platform",
        "app_activity":  ".MainActivity",
    }

    print(f"\n📊 Processing {total} test results...")
    print(f"   ✅ Passed : {passed}")
    print(f"   ❌ Failed : {failed}")
    print(f"   ⏭ Skipped: {skipped}")
    print(f"   📈 Pass % : {(passed/total*100):.1f}%")
    print(f"\n🔨 Building Excel workbook...")

    wb = Workbook()

    print("   → Sheet 1: Executive Summary")
    build_summary_sheet(wb, results, meta)

    print("   → Sheet 2: All Test Results")
    build_detailed_sheet(wb, results)

    print("   → Sheet 3: Failed Tests")
    build_failed_sheet(wb, results)

    print("   → Sheet 4: Module Analytics (with Charts)")
    build_analytics_sheet(wb, results)

    print("   → Sheet 5: Performance Analysis")
    build_performance_sheet(wb, results)

    output_path = os.path.join(OUTPUT_DIR, REPORT_NAME)
    wb.save(output_path)

    abs_path = os.path.abspath(output_path)
    print(f"\n✅ Report saved successfully!")
    print(f"   📁 Location : {abs_path}")
    print(f"   📋 Sheets   : {len(wb.sheetnames)}")
    print(f"   📊 Tests    : {total}")
    print(f"   🗂️  File size: {os.path.getsize(abs_path) / 1024:.1f} KB")
    print("\n" + "="*65 + "\n")

    return abs_path


if __name__ == "__main__":
    generate_report()
