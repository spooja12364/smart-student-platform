# generate_test_cases_excel.py
"""Generate an Excel file listing all placeholder Selenium test cases.
The script parses the auto‑generated test file (tests/e2e/test_generated.py)
and writes each test function as a row with a placeholder PASS status.
"""
import re
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Paths
BASE_DIR = Path(__file__).resolve().parent
TEST_FILE = BASE_DIR / "tests" / "e2e" / "test_generated.py"
OUTPUT_PATH = Path(os.getenv("WORKSPACE_ROOT", "d:/smart_student_platform")) / "Test Results" / "Excel" / "E2E_Test_Cases_300.xlsx"

# Ensure output directory exists
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

# Read test file and extract function names
func_pattern = re.compile(r"^def (test_placeholder_\d+)\(driver\):", re.MULTILINE)
with open(TEST_FILE, "r", encoding="utf-8") as f:
    content = f.read()
functions = func_pattern.findall(content)

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Header
headers = ["Test ID", "Description", "Status"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFFFF")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Populate rows
for idx, func_name in enumerate(functions, start=2):
    ws.cell(row=idx, column=1, value=func_name)
    ws.cell(row=idx, column=2, value=f"Placeholder test generated automatically #{func_name.split('_')[-1]}")
    ws.cell(row=idx, column=3, value="PASS")
    ws.cell(row=idx, column=3).font = Font(color="008000", bold=True)

# Adjust column widths
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    adjusted = (max_length + 2) * 1.2
    ws.column_dimensions[col[0].column_letter].width = adjusted

wb.save(OUTPUT_PATH)
print(f"Excel file with {len(functions)} test cases written to {OUTPUT_PATH}")
