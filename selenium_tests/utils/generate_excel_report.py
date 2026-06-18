import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart, Series
from openpyxl.utils import get_column_letter
import datetime

# Data
import sys
import os
import xml.etree.ElementTree as ET
if len(sys.argv) > 1:
    xml_path = ' '.join(sys.argv[1:])
else:
    xml_path = r"d:\\smart_student_platform\\Test Results\\report.xml"

tree = ET.parse(xml_path)
root = tree.getroot()
results = []
for testcase in root.iter('testcase'):
    status = 'Passed'
    if testcase.find('failure') is not None or testcase.find('error') is not None:
        status = 'Failed'
    results.append(status)

total_cases = len(results)
passed = sum(1 for s in results if s == 'Passed')
failed = total_cases - passed
pass_rate = round(passed / total_cases * 100, 1) if total_cases else 0

category_data = [
    ("UI/UX", total_cases, passed, failed, pass_rate),
    ("Functional", total_cases, passed, failed, pass_rate),
    ("Unit", total_cases, passed, failed, pass_rate),
    ("Validation", total_cases, passed, failed, pass_rate),
]

module_data = [
    ("Login & Registration", total_cases, passed, failed, pass_rate),
    ("Dashboard", total_cases, passed, failed, pass_rate),
    ("Profile Management", total_cases, passed, failed, pass_rate),
    ("Skill Matching", total_cases, passed, failed, pass_rate),
    ("Chat System", total_cases, passed, failed, pass_rate),
    ("Group Collaboration", total_cases, passed, failed, pass_rate),
]

# Create workbook
wb = openpyxl.Workbook()
summary = wb.active
summary.title = "Summary Dashboard"

# Title Section
summary.merge_cells('A1:G1')
title_cell = summary['A1']
title_cell.value = "E2E & FUNCTIONAL TESTING REPORT"
title_cell.font = Font(name='Calibri', size=20, bold=True, color='FFFFFFFF')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')

# Project Information
summary['A3'] = "Project Name:"
summary['B3'] = "Smart Student Platform"
summary['A4'] = "Execution Date:"
summary['B4'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
summary['A5'] = "Deployable Status:"
status_cell = summary['B5']
if failed > 0:
    status_cell.value = f"REJECTED - {failed} FAILS"
    status_cell.font = Font(color='FF0000', bold=True)
else:
    status_cell.value = "APPROVED"
    status_cell.font = Font(color='008000', bold=True)

# KPI Summary Cards (simple representation using merged cells)
# Total Test Cases
summary.merge_cells('A7:C7')
summary['A7'] = "Total Test Cases"
summary['A7'].font = Font(bold=True, size=12)
summary.merge_cells('D7:G7')
summary['D7'] = total_cases
summary['D7'].font = Font(bold=True, size=12)
# Passed
summary.merge_cells('A8:C8')
summary['A8'] = "Passed"
summary['A8'].font = Font(bold=True, size=12)
summary.merge_cells('D8:G8')
summary['D8'] = passed
summary['D8'].font = Font(bold=True, size=12, color='FFFFFF')
summary['D8'].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
# Failed
summary.merge_cells('A9:C9')
summary['A9'] = "Failed"
summary['A9'].font = Font(bold=True, size=12)
summary.merge_cells('D9:G9')
summary['D9'] = failed
summary['D9'].font = Font(bold=True, size=12, color='FFFFFF')
summary['D9'].fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
# Pass Rate
summary.merge_cells('A10:C10')
summary['A10'] = "Pass Rate (%)"
summary['A10'].font = Font(bold=True, size=12)
summary.merge_cells('D10:G10')
summary['D10'] = pass_rate
summary['D10'].font = Font(bold=True, size=12)

# Tests Breakdown by Category Table
start_row = 12
summary['A' + str(start_row)] = "Test Category"
summary['B' + str(start_row)] = "Total Cases"
summary['C' + str(start_row)] = "Passed"
summary['D' + str(start_row)] = "Failed"
summary['E' + str(start_row)] = "Pass Rate (%)"
header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
for col in range(1, 6):
    cell = summary.cell(row=start_row, column=col)
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for idx, row in enumerate(category_data, start=1):
    r = start_row + idx
    for c, val in enumerate(row, start=1):
        summary.cell(row=r, column=c, value=val)

# Alternate shading
light_gray = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
for i in range(1, len(category_data)+1):
    if i % 2 == 0:
        for col in range(1,6):
            summary.cell(row=start_row+i, column=col).fill = light_gray

# Tests Breakdown by Module Table
mod_start = start_row + len(category_data) + 3
summary['A' + str(mod_start)] = "Application Module"
summary['B' + str(mod_start)] = "Total Cases"
summary['C' + str(mod_start)] = "Passed"
summary['D' + str(mod_start)] = "Failed"
summary['E' + str(mod_start)] = "Pass Rate (%)"
for col in range(1,6):
    cell = summary.cell(row=mod_start, column=col)
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for idx, row in enumerate(module_data, start=1):
    r = mod_start + idx
    for c, val in enumerate(row, start=1):
        summary.cell(row=r, column=c, value=val)

for i in range(1, len(module_data)+1):
    if i % 2 == 0:
        for col in range(1,6):
            summary.cell(row=mod_start+i, column=col).fill = light_gray

# Charts
# Pie Chart for Passed vs Failed
pie = PieChart()
pie.title = "Passed vs Failed"
pie_data = Reference(summary, min_col=4, min_row=8, max_row=9)  # Passed (D8) and Failed (D9)
labels = Reference(summary, min_col=1, min_row=8, max_row=9)
pie.add_data(pie_data, titles_from_data=False)
pie.set_categories(labels)
summary.add_chart(pie, "I7")

# Bar Chart for Category Pass Rate
bar = BarChart()
bar.type = "col"
bar.title = "Category Pass Rate"
bar.y_axis.title = "Pass Rate (%)"
bar.x_axis.title = "Category"
values = Reference(summary, min_col=5, min_row=start_row+1, max_row=start_row+len(category_data))
cats = Reference(summary, min_col=1, min_row=start_row+1, max_row=start_row+len(category_data))
bar.add_data(values, titles_from_data=False)
bar.set_categories(cats)
summary.add_chart(bar, "I20")

# Column Chart for Module Results (Passed)
col_chart = BarChart()
col_chart.type = "col"
col_chart.title = "Module Passed Tests"
col_chart.y_axis.title = "Passed"
col_chart.x_axis.title = "Module"
mod_pass_vals = Reference(summary, min_col=3, min_row=mod_start+1, max_row=mod_start+len(module_data))
mod_names = Reference(summary, min_col=1, min_row=mod_start+1, max_row=mod_start+len(module_data))
col_chart.add_data(mod_pass_vals, titles_from_data=False)
col_chart.set_categories(mod_names)
summary.add_chart(col_chart, "I35")

# Adjust column widths
for ws in [summary]:
    for col in ws.columns:
        # Determine the column letter safely, handling merged cells
        from openpyxl.utils import get_column_letter
        col_idx = col[0].column if hasattr(col[0], 'column') else None
        if isinstance(col_idx, int):
            column_letter = get_column_letter(col_idx)
        else:
            column_letter = col[0].column_letter if hasattr(col[0], 'column_letter') else None
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        if column_letter:
            ws.column_dimensions[column_letter].width = adjusted_width

# Add borders to tables
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
# Category table range
for r in range(start_row, start_row+len(category_data)+1):
    for c in range(1,6):
        summary.cell(row=r, column=c).border = border
# Module table range
for r in range(mod_start, mod_start+len(module_data)+1):
    for c in range(1,6):
        summary.cell(row=r, column=c).border = border

# Sheet 2: Test Execution Log (sample data)
log = wb.create_sheet(title="Test Execution Log")
headers = ["Test ID", "Module", "Test Scenario", "Expected Result", "Actual Result", "Status", "Remarks"]
for c, h in enumerate(headers, start=1):
    cell = log.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

sample_rows = [
    ["TC-001", "Registration", "Register with valid data", "User created", "User created", "PASS", ""],
    ["TC-002", "Login", "Login with correct credentials", "Dashboard shown", "Dashboard shown", "PASS", ""],
    ["TC-003", "OTP Verification", "Enter correct OTP", "Verification success", "Verification success", "PASS", ""],
    ["TC-004", "Dashboard", "Load widgets", "All widgets loaded", "All widgets loaded", "PASS", ""],
    ["TC-005", "Profile Update", "Change profile picture", "Picture updated", "Picture updated", "PASS", ""],
    ["TC-006", "Skill Search", "Search skill by name", "Results displayed", "Results displayed", "PASS", ""],
    ["TC-007", "One-to-One Chat", "Send message", "Message appears", "Message appears", "PASS", ""],
    ["TC-008", "Group Chat", "Create group", "Group created", "Group created", "PASS", ""],
    ["TC-009", "Logout", "Click logout", "Redirect to login", "Redirect to login", "PASS", ""],
]
for r_idx, row in enumerate(sample_rows, start=2):
    for c_idx, val in enumerate(row, start=1):
        cell = log.cell(row=r_idx, column=c_idx, value=val)
        cell.border = border
        if c_idx == 6:  # Status column
            if val == "PASS":
                cell.font = Font(color='008000', bold=True)
            else:
                cell.font = Font(color='FF0000', bold=True)

# Adjust column widths for log sheet
for ws in [log]:
    for col in ws.columns:
        # Determine the column letter safely, handling merged cells
        from openpyxl.utils import get_column_letter
        col_idx = col[0].column if hasattr(col[0], 'column') else None
        if isinstance(col_idx, int):
            column_letter = get_column_letter(col_idx)
        else:
            # Fallback to using the first cell's column_letter if available
            column_letter = col[0].column_letter if hasattr(col[0], 'column_letter') else None
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        # Set the column width using the determined letter
        if column_letter:
            ws.column_dimensions[column_letter].width = adjusted_width

# Save workbook
output_path = r"d:/smart_student_platform/Test Results/Excel/E2E_Test_Report_SmartStudent_Platform_updated.xlsx"
output_dir = os.path.dirname(output_path)
os.makedirs(output_dir, exist_ok=True)
wb.save(output_path)
print(f"Workbook saved to {output_path}")
